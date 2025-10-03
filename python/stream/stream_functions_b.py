import base64
import json
import os
import tempfile
import threading
import time
import zipfile
import numpy
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests as r
import solara
from ipydatagrid import DataGrid
from ipyleaflet import GeoJSON, LegendControl, Map
from ipystream.renderer import plotly_fig_to_html
from ipywidgets import HTML, Button, RadioButtons, widgets
from openpyxl.utils import get_column_letter
from pandas import DataFrame
from ipystream.stream import Stream, WidgetCurrentsChildren
from python.login import headers

col_value = "Capacity (kW or KWh)"
col_value_net = "Capacity (kW)"
filters_that_apply = ["Solution", "Stage"]
building_layer = "Plan_guide_V2 - adapted"
hovered_net_color = "purple"
#nelly quick-fix netowkrs
network_to_color = {"CAD": "red", 
                    "AAD": "#FFED29", 
                    #"other": "yellow" # Mittlenetz
                    #"other": "blue" # Cooling
                    "other": "red" # Thermal
                   }
skip_last_level_key = "skip_last_level"
bar_path = f"iframe_figures/bar{str(time.time()).replace('.','_')}.html"


def main_stream(project_id):
    # init
    cache = {}
    base_url = "https://eu-north-1-api.sympheny.com/"
    cache["base_url"] = base_url
    cache["be"] = f"{base_url}sympheny-app/"
    headers(cache)
    cache["execution_name"] = "exec"
    cache["col_name"] = "Name"
    cache["sleep"] = 5
    cache["col_name"] = "Name"
    cache["skip_download_results"] = True

    cache["col_value"] = col_value
    s = Stream(cache=cache, debounce_sec=0.5)
    s.register(
        1, widgets=[lambda s, p=project_id: select_scenario(p, s)], title="1) Select scenario (scenarios without results are ignored)"
    )
    s.register(2, updater=select_result_UPDATE, title="2) Select execution result")
    s.register(3, updater=filter_widgets_UPDATE, title="3) Filter results (Keep CTRL pressed to select multiple)")
    s.register(4, updater=df_table_UPDATE, vertical=True, title="4) Display results")
    s.display_registered()
    cache["stream"] = s


# LEVEL 1
def select_scenario(project_id, s: Stream):
    cache = s.cache
    cache["project_id"] = project_id
    base_url = cache["base_url"]
    be = cache["be"]
    h = cache["h"]

    scenario_display_to_guids = {}

    no_execution_results = []
    analyses = r.get(f"{be}projects/{project_id}", headers=h).json()["data"]["analyses"]
    for a in analyses:
        a_name = a["analysisName"]
        a_guid = a["analysisGuid"]

        for s in a["scenarios"]:
            s_name = s["scenarioName"]
            s_guid = s["scenarioGuid"]
            option = f"Analys: {a_name} | Scenar: {s_name}"

            done_jobs = get_done_jobs(s_guid, base_url, h)
            if done_jobs:
                scenario_display_to_guids[option] = (a_guid, a_name, s_guid, s_name)
            else:
                no_execution_results.append(option)

    cache["scenario_display_to_guids"] = scenario_display_to_guids

    options = list(scenario_display_to_guids)
    options = [(x, x) for x in options]
    options.sort()

    return RadioButtons(
        options=options,
        layout={"width": "max-content"},
        disabled=False,
    )

    # no_execution_results.sort()
    # print("-- Following scenarios are ignored, because no execution results: \n")
    # print(no_execution_results)


def get_done_jobs(scenario_id, base_url, h):
    data = {"scenarioGuids": [scenario_id], "limit": 200}
    resp = r.post(f"{base_url}sense-api/ext/solver/jobs/get-scenarios", headers=h, json=data).json()
    return list(filter(lambda j: j["scenarioGuid"] == scenario_id and j["status"] == "DONE", resp))


# LEVEL 2
def select_result_UPDATE(w: WidgetCurrentsChildren):
    cache = w.cache
    base_url = cache["base_url"]
    h = cache["h"]

    dropdown = w.parents[0]
    selected = dropdown.value
    (analysis_id, analysis_name, scenario_id, scenario_name) = cache["scenario_display_to_guids"][selected]
    done_jobs = get_done_jobs(scenario_id, base_url, h)
    started_to_job_id = {f'{x["started"]} | {x["name"]}': x["id"] for x in done_jobs}

    cache["started_to_job_id"] = started_to_job_id
    cache["analysis_id"] = analysis_id
    cache["analysis_name"] = analysis_name
    cache["scenario_id"] = scenario_id
    cache["scenario_name"] = scenario_name

    opts = [(x, x) for x in started_to_job_id]

    # update widget
    widg = RadioButtons(layout={"width": "max-content"}, options=opts, value=opts[0][0])
    w.display_or_update(widg)


# LEVEL 3
filter_cols = ["Solution", "Stage", "Type", "Hub or Link"]

niveau_to_name = {
    "CAD": ["Echangeur CAD"],
    "Secteur": ["PAC Air-Eau (SECT - ECS)", "PAC Eau-Eau (SECT - ECS)"],
    "Réseau secteur": ["Réseau SECT (Anergy)", "Réseau SECT (ECS/CH)"],
    "Bâtiment": [
        "Booster Eau-Eau (BAT - ECS)",
        "PAC Air-Eau (BAT - CH/ECS)",
        "PAC Eau-Eau (BAT - CH/ECS)",
        "Groupe froid (BAT)",
        "Free Cooling",
    ],
    "Réseau": ["CAD existant"],
    "Import": ["AAD"],
    "Dummy": [],
}


def select_multi_widget(desc, opts):
    if desc == "Hub or Link":
        desc = "Hub"

    value = opts
    if desc == "Solution":
        value = [opts[0]]

    return widgets.SelectMultiple(description=desc, options=opts, value=value)


def build_col_to_widget(widgets):
    col_to_widget = {}
    for i, col in enumerate(filter_cols):
        col_to_widget[col] = widgets[i]
    return col_to_widget


def filter_widgets_UPDATE(w: WidgetCurrentsChildren):
    cache = w.cache
    dropdown_job = w.parents[0]
    build_df(dropdown_job, w.cache)
    df = cache["df"]
    df_copy = df.copy()
    df_copy.drop(df_copy[df_copy["Type"] == "Network"].index, inplace=True)

    col_to_opts = {}
    for col in filter_cols:
        opts = list(set(df_copy[col]))
        opts.sort()
        col_to_opts[col] = opts

    # w_niveau
    col_name = cache["col_name"]
    names = [x for xs in niveau_to_name.values() for x in xs]
    for name in df[col_name]:
        if name not in names:
            niveau_to_name["Dummy"].append(name)
    cache["niveau_to_name"] = niveau_to_name

    opts_niveau = list(niveau_to_name.keys())
    opts_niveau.sort()

    # update widgets
    for col, opts in col_to_opts.items():
        widg = select_multi_widget(col, opts)
        w.display_or_update(widg)

    widg_niveau = select_multi_widget("Niveau prod", opts_niveau)
    w.display_or_update(widg_niveau)


def apply_filter(col_to_widget, df):
    if df.empty:
        return df
    
    col_to_filter = {k: list(v.value) for k, v in col_to_widget.items() if v.value}
    for col, filt in col_to_filter.items():
        df = df[df[col].map(lambda x, f=filt: x in f)]

    return df


def remove_virtual(df):
    col = "Lifetime (years)"
    return df[df[col] != 1]


def download_results(scenario, analysis, execution_name, job_id, headers, cache):
    response = r.get(f"https://eu-north-1-api.sympheny.com/sense-api/ext/solver/jobs/{job_id}", headers=headers)
    job_name = response.json()["name"]

    input_dir = f"results_repository/{analysis}/{scenario}"
    input_path = f"{input_dir}/{execution_name}-{scenario}.xlsx"
    result_dir = f"results_repository/{analysis}/{scenario}/{job_name}/{job_id}"
    result_path = f"{result_dir}/{scenario}.zip"

    if not os.path.exists(input_dir):
        os.makedirs(input_dir)

    if not os.path.exists(result_dir):
        os.makedirs(result_dir)
    elif cache.get("skip_download_results", False):
        return result_dir

    input_file = response.json()["inputFile"]
    output_file = response.json()["outputFile"]

    response = r.get(input_file, stream=True)
    with open(input_path, "wb") as f:
        f.write(response.content)

    response = r.get(output_file, stream=True)
    with open(result_path, "wb") as f:
        f.write(response.content)

    try:
        with zipfile.ZipFile(result_path, "r") as zip_ref:
            zip_ref.extractall(result_dir)
    except Exception as e:
        print("Failed unzip results: ", e)

    return result_dir


def build_df(dropdown_job, cache):
    h = cache["h"]
    started_to_job_id = cache["started_to_job_id"]
    analysis_name = cache["analysis_name"]
    scenario_name = cache["scenario_name"]
    execution_name = cache["execution_name"]
    selected_job = dropdown_job.value
    job_id = started_to_job_id[selected_job]
    cache["job_id"] = job_id
    result_dir = download_results(scenario_name, analysis_name, execution_name, job_id, h, cache)
    cache["result_dir"] = result_dir

    df, df_network = read_excel(result_dir)
    cache["df"] = df
    #nelly quick-fix netowkrs: modify the dataframe to only allow one network
    #df_network = df_network[df_network["Name"] == "Mitteltemperaturnetz"]
    #df_network = df_network[df_network["Name"] == "Cooling Network"]
    #df_network = df_network[df_network["Name"] == "Thermal network 65-40°C"]

    cache["df_network"] = df_network


def read_excel(result_dir):
    excel_name = [x for x in os.listdir(result_dir) if x.startswith("Summary")][0]
    excel = f"{result_dir}/{excel_name}"
    dfs = pd.read_excel(excel, sheet_name=["Cost & CO2", "Networks"])
    return dfs["Cost & CO2"], dfs["Networks"]


# LEVEL 4a
def df_table_UPDATE(w: WidgetCurrentsChildren):
    cache = w.cache
    df = cache["df"]
    col_to_widget = build_col_to_widget(w.parents)
    w_niveau = w.parents[-1]
    niveau_to_name = cache["niveau_to_name"]
    col_name = cache["col_name"]

    df_filt = df.copy()
    df_filt = apply_filter(col_to_widget, df_filt)
    df_filt = df_filt.map(format_thousands)
    df_filt = df_filt.dropna(axis=1, how="all")

    niveau_selected = w_niveau.value
    if niveau_selected:
        niveau_selected_names = [l for k, l in niveau_to_name.items() if k in niveau_selected]
        niveau_selected_names = [x for xs in niveau_selected_names for x in xs]
        df_filt = df_filt[df_filt[col_name].map(lambda x, f=niveau_selected_names: x in f)]

    # for tests
    cache["df_filt"] = df_filt
    if skip_last_level_key in cache and cache[skip_last_level_key]:
        return

    # update widget
    grid = DataGrid(
        df_filt,
        selection_mode="cell",
        base_column_size=200,
        base_row_header_size=300,
        layout={"height": "250px"},
    )
    grid.auto_fit_columns = True

    w.sub_title("SANKEY URL (using first value of Solution, Stage and Hub dropdowns)")
    sankey_url(w)

    w.sub_title("Results table")
    w.display_or_update(widgets.VBox([grid]))
    w.display_or_update(download_button(df_filt))

    w.sub_title("Aggregate on (Hub)")
    df_table_agg(w)

    chart_pie(w)
    chart_bar_vert(w)
    chart_bar(w)
    chart_download(w)

    w.sub_title("Network map (only sensitive to Solution and Stage filters)")
    display_NETWORK(w)


def format_thousands(x):
    if isinstance(x, float) and not numpy.isnan(x):
        x = round(x)
        apostrophe = "\u0027"
        return f"{x:,}".replace(",", apostrophe)

    return x


def sankey_url(w: WidgetCurrentsChildren):
    cache = w.cache
    col_to_widget = build_col_to_widget(w.parents)

    if not col_to_widget["Solution"].value or not col_to_widget["Stage"].value or not col_to_widget["Hub or Link"].value:
        sankey_url = "SELECT AT LEAST 1 Solution, 1 Stage and 1 Hub"
    else:
        point = col_to_widget["Solution"].value[0].split(" ")[1]
        stage = col_to_widget["Stage"].value[0]
        hub = col_to_widget["Hub or Link"].value[0]

        project_id = cache["project_id"]
        analysis_id = cache["analysis_id"]
        job_id = cache["job_id"]

        sankey_url = (
            f"https://app.sympheny.com/projects/{project_id}/analysis/{analysis_id}"
            f"/execution/{job_id}/solution/{point}/general?stage={stage}&hub={hub}"
        )
        sankey_url = sankey_url.replace(" ", "%20")
        sankey_url = f'<a style="color: blue;" target="_blank" rel="noopener noreferrer" href="{sankey_url}">{sankey_url}</a>'

    w.display_or_update(HTML(sankey_url))


html_template = """<html>
<head>
<meta name="viewport" content="width=device-width, initial-scale=1">
</head>
<body>
<a download="{filename}" href="data:text/csv;base64,{payload}" download>
<button class="p-Widget jupyter-widgets jupyter-button widget-button mod-warning">Download File</button>
</a>
</body>
</html>
"""


def download_button(df: DataFrame) -> HTML:
    sheet_name = "0"
    bytes = None
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
        path = f.name

        # adjust columns size
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            worksheet = writer.sheets[sheet_name]
    
            for column_cells in worksheet.iter_cols():
                if column_cells:
                    col_idx = column_cells[0].column 
                    new_column_letter = get_column_letter(col_idx)
    
                    header_text = str(df.columns[col_idx - 1]) if col_idx - 1 < len(df.columns) else ""
                    
                    max_length = len(header_text)
                    for cell in column_cells:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
    
                    adjusted_width = max_length * 1.23
                    if adjusted_width > 0:
                        worksheet.column_dimensions[new_column_letter].width = adjusted_width

        with open(path, "rb") as file:
            bytes = file.read()

    payload = base64.b64encode(bytes).decode()
    html_button = html_template.format(payload=payload, filename="results.xlsx")
    return HTML(html_button)


# AGGREGATE
def df_table_agg(w: WidgetCurrentsChildren):
    col_aggrs = ["Hub or Link"]
    col_drops = ["Solution", "Stage", "Type", "Name", "Hub or Link"]

    cache = w.cache
    df = cache["df"]
    col_to_widget = build_col_to_widget(w.parents)
    w_niveau = w.parents[-1]
    col_name = cache["col_name"]

    df_filt = df.copy()
    df_filt = apply_filter(col_to_widget, df_filt)

    niveau_selected = w_niveau.value
    if niveau_selected:
        niveau_selected_names = [l for k, l in niveau_to_name.items() if k in niveau_selected]
        niveau_selected_names = [x for xs in niveau_selected_names for x in xs]
        df_filt = df_filt[df_filt[col_name].map(lambda x, f=niveau_selected_names: x in f)]

    df_filt = df_table_agg_filt(df_filt, col_aggrs, col_drops)

    grid = DataGrid(
        df_filt,
        selection_mode="cell",
        base_column_size=200,
        base_row_header_size=300,
        layout={"height": "250px"},
    )
    grid.auto_fit_columns = True

    w.display_or_update(widgets.VBox([grid]))
    w.display_or_update(download_button(df_filt))


def df_table_agg_filt(df_filt, col_aggrs, col_drops):
    col_drops = [x for x in col_drops if x in df_filt.columns and x not in col_aggrs]
    df_filt = df_filt.drop(columns=col_drops)

    df_filt = df_filt.dropna(axis=1, how="all")

    if not set(col_aggrs).issubset(df_filt.columns):
        return pd.DataFrame({col_aggrs[0]: []})

    df_filt = df_filt.groupby(col_aggrs, as_index=False).sum()
    df_filt = df_filt.map(format_thousands)
    return df_filt


# LEVEL 4b
def chart_pie(w: WidgetCurrentsChildren):
    cache = w.cache
    df = cache["df"]
    col_to_widget = build_col_to_widget(w.parents)
    w_niveau = w.parents[-1]
    col_name = cache["col_name"]
    col_value = cache["col_value"]

    df_filt = df.copy()
    df_filt = df_filt[df_filt[col_value] > 0]
    df_filt = apply_filter(col_to_widget, df_filt)
    df_filt = remove_virtual(df_filt)

    niveau_selected = w_niveau.value
    if niveau_selected:
        niveau_selected_names = [l for k, l in niveau_to_name.items() if k in niveau_selected]
        niveau_selected_names = [x for xs in niveau_selected_names for x in xs]
        df_filt = df_filt[df_filt[col_name].map(lambda x, f=niveau_selected_names: x in f)]

    df_filt = df_filt[df_filt.columns.intersection([col_name, col_value])]
    df_filt = df_filt.groupby(col_name, as_index=False).sum()

    # draw
    names = df_filt[col_name]
    values = [round(x, 2) for x in df_filt[col_value]]
    legend_labels = [f"{format_thousands(x)} k" for x in values]

    # pie chart
    fig1 = go.Figure(data=[go.Pie(labels=names, values=values, textinfo="text+percent", hoverinfo="label", text=legend_labels)])
    fig1.update_traces(textposition="inside")
    fig1.update_layout(title=col_value, legend_title=col_name, width=800, height=800, uniformtext_minsize=12, uniformtext_mode="hide")
    w.display_or_update(plotly_fig_to_html(fig1))


# LEVEL 4c
def chart_bar_vert(w: WidgetCurrentsChildren):
    cache = w.cache
    df = cache["df"]
    col_to_widget = build_col_to_widget(w.parents)
    w_niveau = w.parents[-1]
    col_name = cache["col_name"]
    col_value = cache["col_value"]

    df_filt = df.copy()
    df_filt = df_filt[df_filt[col_value] > 0]
    df_filt = apply_filter(col_to_widget, df_filt)
    df_filt = remove_virtual(df_filt)

    niveau_selected = w_niveau.value
    if niveau_selected:
        niveau_selected_names = [l for k, l in niveau_to_name.items() if k in niveau_selected]
        niveau_selected_names = [x for xs in niveau_selected_names for x in xs]
        df_filt = df_filt[df_filt[col_name].map(lambda x, f=niveau_selected_names: x in f)]

    df_filt = df_filt[df_filt.columns.intersection([col_name, col_value])]
    df_filt = df_filt.groupby(col_name, as_index=False).sum()
    names = df_filt[col_name]
    values = [round(x, 2) for x in df_filt[col_value]]
    name_values = list(zip(names, values))
    name_values = list(sorted(name_values, key=lambda x: x[1], reverse=True))

    # name_values = name_values[:20]
    names = [n for n, _ in name_values]
    values = [v for _, v in name_values]

    values = [round(v, 2) for v in values]
    data = {"name": names, "value": values}
    data = pd.DataFrame.from_dict(data)

    fig = px.bar(data, x="name", y="value", text="value")
    fig.update_traces(textposition="none")
    fig.update_layout(xaxis_title="", yaxis_title=col_value, legend_traceorder="normal")

    for i, _ in enumerate(fig.data):
        fig.data[i]["hovertemplate"] = "%{y:,.0f}"

    w.display_or_update(plotly_fig_to_html(fig))


# LEVEL 4d
def chart_bar(w: WidgetCurrentsChildren):
    cache = w.cache
    df = cache["df"]
    col_to_widget = build_col_to_widget(w.parents)
    w_niveau = w.parents[-1]
    col_name = cache["col_name"]
    col_value = cache["col_value"]

    df_filt = df.copy()
    df_filt = df_filt[df_filt[col_value] > 0]
    df_filt = apply_filter(col_to_widget, df_filt)
    df_filt = remove_virtual(df_filt)

    niveau_selected = w_niveau.value
    if niveau_selected:
        niveau_selected_names = [l for k, l in niveau_to_name.items() if k in niveau_selected]
        niveau_selected_names = [x for xs in niveau_selected_names for x in xs]
        df_filt = df_filt[df_filt[col_name].map(lambda x, f=niveau_selected_names: x in f)]

    df_filt = df_filt[df_filt.columns.intersection([col_name, col_value])]
    df_filt = df_filt.groupby(col_name, as_index=False).sum()
    names = df_filt[col_name]
    values = [round(x, 2) for x in df_filt[col_value]]
    name_values = list(zip(names, values))
    name_values = list(sorted(name_values, key=lambda x: x[1], reverse=True))
    total = sum([v for _, v in name_values])

    # name_values = name_values[:20]
    names = [n for n, _ in name_values]
    values = [v for _, v in name_values]
    y = [0] * len(names)
    labels = [f"{round(v*100/total, 2)} %" for v in values]
    data = {"name": names, "value": values, "label": labels, "y": y}
    data = pd.DataFrame.from_dict(data)

    fig = px.bar(data, x="value", y="y", color="name", labels="name", text="label", orientation="h")
    fig.update_yaxes(visible=False, showticklabels=False)
    fig.update_traces(textposition="inside")
    fig.update_layout(barmode="stack", legend_traceorder="normal", uniformtext_minsize=12, uniformtext_mode="hide", xaxis_title=col_value)

    for i, trace in enumerate(fig.data):
        fig.data[i]["hovertemplate"] = trace.name

    fig_widget = plotly_fig_to_html(fig)
    fig_html = fig_widget.value
    with open(bar_path, "w", encoding="utf-8") as f:
        f.write(fig_html)

    w.display_or_update(fig_widget)


# LEVEL 4e
def chart_download(w: WidgetCurrentsChildren):
    download_html = widgets.VBox([HTML("")])
    w.display_or_update(download_html)

    def on_done():
        download_html.children = [HTML("")]

    def chart():
        download_html.children = [HTML("<font size='3' color='red'>Downloading, please wait ...</font>")]

        with open(bar_path, "rb") as f:
            res = f.read()
            threading.Timer(1, on_done).start()
            return res

    dl = solara.FileDownload(chart, filename="bar.html", label="Download chart")
    w.display_or_update(dl)


# LEVEL 4f
def display_NETWORK(w: WidgetCurrentsChildren):
    cache = w.cache
    button = Button(button_style="danger", description="Loading....", disabled=True)
    button_box = widgets.VBox([button])
    w.display_or_update(button_box)

    network_text_empty = hovered_network_text(None, None)
    network_test_key = "network_text"
    network_text = cache.get(network_test_key)
    if not network_text:
        network_text = network_text_empty
        cache[network_test_key] = network_text
    else:
        network_text.children = network_text_empty.children

    w.display_or_update(network_text)

    # build map
    m = build_map_NETWORK(w, network_text)
    w.display_or_update(widgets.VBox([m]))

    # re-enable center map button
    button.description = "Center map"
    button.disabled = False
    button_box.children = [button]

    # "center map" on_button_clicked
    def on_button_clicked(_):
        bounds = cache["bounds"]
        m.fit_bounds(bounds)

    button.on_click(on_button_clicked)


def build_map_NETWORK(w: WidgetCurrentsChildren, network_text):
    cache = w.cache
    df_net = cache["df_network"]
    
    # filter dataframe
    col_to_widget = build_col_to_widget(w.parents)
    col_to_widget = {k: v for k, v in col_to_widget.items() if k in filters_that_apply}
    df_net = apply_filter(col_to_widget, df_net)

    # aggregate duplicates
    df_net = df_net.groupby(["Link"], as_index=False).sum()

    # get links
    networks = [(row["Link"], row[col_value_net]) for _, row in df_net.iterrows()]

    be = cache["be"]
    h = cache["h"]
    base_url = cache["base_url"]
    scenario_id = cache["scenario_id"]

    network_layers = r.get(f"{base_url}api-services/gis/scenarios/{scenario_id}/networks", headers=h).json()
    link_id_to_geojson = {x["link_id"]: {"features": [x["feature"]]} for x in network_layers}
    scenario_links = r.get(f"{be}v2/scenarios/{scenario_id}/network-links", headers=h).json()["data"]
    link_id_to_name = {x["networkLinkGuid"]: x["name"] for x in scenario_links}
    link_name_to_id = {v: k for k, v in link_id_to_name.items()}
    link_ids = [x for x in link_id_to_name.keys() if x in link_id_to_geojson]
    link_to_geojson = {link_id_to_name[id]: link_id_to_geojson[id] for id in link_ids}

    network_geojsons = []
    missing_link_geojson = []
    link_id_to_value = {}
    for net in networks:
        link_name = net[0]
        link_value = net[1]
        if link_name not in link_to_geojson:
            missing_link_geojson.append(link_name)
            continue

        geojson = link_to_geojson[link_name]
        network_geojsons.append((link_value, geojson, link_name))

        link_id = link_name_to_id[link_name]
        link_id_to_value[link_id] = link_value

    # build map
    max_value = max([x[0] for x in network_geojsons]) if network_geojsons else 0
    m = Map()
    m.layout.width = "100%"
    m.layout.height = "500px"
    m.add(
        LegendControl(
            {
                "Network CAD": network_to_color["CAD"],
                "Network AAD": network_to_color["AAD"],
                "Network other": network_to_color["other"],
                f"Layer {building_layer}": "green",
                "hubs": "lightblue",
            },
            name=f"Thickest network: {format_thousands(max_value)} kW",
            position="topright",
        )
    )

    # add building layer 'Plan_guide_V2 - adapted'
    layers = r.get(f"{base_url}api-services/gis/scenarios/{scenario_id}/layers-presigned", headers=h).json()
    layers_filtered = [x["layer_id"] for x in layers if x["layer_name"] == building_layer]

    if layers_filtered:
        layer_id = layers_filtered[0]
        layer_url = r.get(f"{base_url}api-services/gis/scenarios/{scenario_id}/layers-presigned/{layer_id}", headers=h).json()["url"]
        geojson = json.loads(r.get(layer_url).content.decode("utf-8"))["feature_collection"]
        m.add(GeoJSON(data=geojson, style={"color": "green", "weight": 1}))

    # add hubs
    resp = r.get(f"{be}scenarios/{scenario_id}/hubs", headers=h).json()["data"]
    hub_ids = [x["hubGuid"] for x in resp]

    for hub_id in hub_ids:
        resp = r.get(f"{base_url}api-services/gis/scenarios/{scenario_id}/hubs/{hub_id}", headers=h).json()
        if not resp:
            continue

        geojson_base = resp["base_layer"]
        m.add(GeoJSON(data=geojson_base, style={"color": "deepskyblue", "weight": 2}))

    # add networks
    def on_hover_network(*args, **kwargs):
        link_id = kwargs["feature"]["properties"]["link_id"]
        name = link_id_to_name[link_id]
        value = link_id_to_value[link_id]

        network_text_updated = hovered_network_text(name, value)
        network_text.children = network_text_updated.children

    features = []
    for value, geojson, name in network_geojsons:
        weight = 15 * value / max_value
        features.append(geojson["features"][0])
        hover_style = {"color": hovered_net_color, "dashArray": "0", "fillOpacity": 0.5}
        feat = GeoJSON(data=geojson, style={"color": net_color(name), "weight": weight}, hover_style=hover_style)

        feat.on_hover(on_hover_network)

        m.add(feat)

    cache["bounds"] = get_bounds({"features": features})
    return m


def get_bounds(geojson):
    minX, minY, maxX, maxY = None, None, None, None
    for ft in geojson["features"]:
        coords = ft["geometry"]["coordinates"][0][0]
        if not isinstance(coords, list):
            coords = ft["geometry"]["coordinates"]

        [x0, y0] = coords[0]
        if not minX:
            minX = x0
            maxX = x0
            minY = y0
            maxY = y0

        for c in coords:
            [x, y] = c
            if x < minX:
                minX = x
            if x > maxX:
                maxX = x
            if y < minY:
                minY = y
            if y > maxY:
                maxY = y

    return [[minY, minX], [maxY, maxX]]


def net_color(network):
    if "CAD" in network:
        return network_to_color["CAD"]
    elif "AAD" in network:
        return network_to_color["AAD"]
    else:
        return network_to_color["other"]


def hovered_network_text(name, value):
    if not name or not value:
        name = "_"
        value = "_"
    else:
        value = format_thousands(value)

    text = (
        f"<font color='black' size=5 style='font-weight: normal'>Hovered network name: </>"
        f"<font size=5 color='{hovered_net_color}' style='font-weight: bold'>{name}</>"
        f"<br></><font color='black' size=5 style='font-weight: normal'>Hovered network value: </>"
        f"<font size=5 color='{hovered_net_color}' style='font-weight: bold'>{value} kW</>"
    )
    return widgets.VBox([HTML(text)])
